#!/usr/bin/env python3
# SARbayes/ISRID/update.py

"""
SARbayes

Database cleaning routine (improved).
"""


import argparse
import datetime
import formatting
import geopy
import openpyxl
import util
import warnings
import weather
import yaml


def update_row(index, input_row, column_settings, output_worksheet, styles):
    row_values = list(input_cell.value for input_cell in input_row)
    if index > 0:
        formatting.standardize(index, row_values, column_settings)
        date, lat, lon = (
            row_values[column_settings['date']['index']], 
            row_values[column_settings['lkp_ns']['index']], 
            row_values[column_settings['lkp_ew']['index']]
        )
        
        snow = row_values[column_settings['snow']['index']]
        if type(snow) is str and 'NO' in snow.upper():
            row_values[column_settings['snow']['index']] = 0.0
        rain = row_values[column_settings['rain']['index']]
        if type(rain) is str and 'NO' in rain.upper():
            row_values[column_settings['rain']['index']] = 0.0
        
        if type(date) is datetime.datetime and \
                type(lat) is float and type(lon):
            coordinates = geopy.Point(lat, lon)
            conditions = weather.get_conditions(date, coordinates)
            
            if type(conditions['TMAX']) is float:
                temp_max = round(conditions['TMAX']/10, 3)
                row_values[column_settings['temp_max']['index']] = temp_max
                util.log('Case {}: Temp/H -> {} deg. C'.format(index, temp_max))
            
            if type(conditions['TMIN']) is float:
                temp_min = round(conditions['TMIN']/10, 3)
                row_values[column_settings['temp_min']['index']] = temp_min
                util.log('Case {}: Temp/L -> {} deg C'.format(index, temp_min))
            
            if type(conditions['AWND']) is float:
                # 1 m/s = 3.6 km/h
                wind_speed = round(conditions['AWND']/1000*60*60, 3)
                row_values[column_settings['wind_speed']['index']] = wind_speed
                util.log('Case {}: Wind Speed -> {} km/h'.format(
                    index, wind_speed))
            
            if type(conditions['SNOW']) is float:
                snow = round(conditions['SNOW'], 3)
                row_values[column_settings['snow']['index']] = snow
                util.log('Case {}: Snowfall -> {} mm'.format(index, snow))
            
            if type(conditions['PRCP']) is float:
                prcp = round(conditions['PRCP'], 3)
                snow = row_values[column_settings['snow']['index']]
                if type(snow) is float:
                    rain = round(max(prcp - snow, 0.0), 3)
                elif type(snow) is float and 'YES' in snow.upper():
                    rain = 0.0
                else:
                    rain = round(prcp, 3)
                row_values[column_settings['rain']['index']] = rain
                util.log('Case {}: Rainfall -> {} mm'.format(index, rain))
    
    for value in row_values:
        output_cell = openpyxl.writer.dump_worksheet.WriteOnlyCell(
            output_worksheet, value)
        for attribute, style in styles.items():
            setattr(output_cell, attribute, style)
        else:
            yield output_cell


def update_database(input_filename, output_filename, settings_filename):
    util.log('Updating database ... ')
    
    util.log('Reading settings file ... ')
    with open(settings_filename) as settings_file:
        settings = yaml.load(settings_file)
        style_settings = settings.get('styles', dict())
        column_settings = settings.get('columns', dict())
        title, styles = settings.get('title', 'ISRID'), dict()
        weather.API_TOKENS = settings['tokens']
        
        util.log('Creating styles ... ')
        for attribute, arguments in style_settings.items():
            if attribute == 'font':
                style = openpyxl.styles.Font(**arguments)
            else:
                style = None
            styles[attribute] = style
    
    util.log('Reading input (optimized) ... ')
    input_workbook = openpyxl.load_workbook(input_filename, read_only=True)
    input_worksheet = input_workbook.active
    
    util.log('Creating output workbook ... ')
    output_workbook = openpyxl.Workbook(write_only=True, optimized_write=True)
    output_worksheet = output_workbook.create_sheet(title=title)
    
    util.log('Writing output (optimized) ... ')
    for index, input_row in enumerate(input_worksheet.rows):
        output_row = update_row(
            index, input_row, column_settings, output_worksheet, styles)
        output_worksheet.append(output_row)
    else:
        output_workbook.save(output_filename)
        util.log('Done.')


def main():
    try:
        parser = argparse.ArgumentParser(description='Update ISRID.')
        parser.add_argument('input_filename', 
            help='Filename of the input spreadsheet.')
        parser.add_argument('output_filename', 
            help='Filename of the output spreadsheet.')
        parser.add_argument('settings_filename', 
            help='Filename of the settings (in YAML format).')
        
        args = parser.parse_args()
        warnings.filterwarnings('ignore')
        update_database(**vars(args))
    except KeyboardInterrupt:
        quit(1)


if __name__ == '__main__':
    main()
