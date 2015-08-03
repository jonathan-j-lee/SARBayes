#!/usr/bin/env python3
# SARbayes/ISRID/integrate.py


import datetime
import geopy
import formatting
import openpyxl
import re
import weather
import yaml
import util

DATETIME = re.compile(r'(\d+)[\-/\.](\d+)[\-/\.](\d+)')
geolocator = geopy.geocoders.GoogleV3(api_key='AIzaSyA8S2jAM6Er6ZsbB3HJ6pjy60NV2cFOcTY')

MONTHS = (
    'JAN', 
    'FEB', 
    'MAR', 
    'APR', 
    'MAY', 
    'JUN', 
    'JUL', 
    'AUG', 
    'SEP', 
    'OCT', 
    'NOV', 
    'DEC'
)

YMD = re.compile(r'(\d{4})[/\s]+(' + r'|'.join(MONTHS) + r').*[/\s]+(\d+)')
DMY = re.compile(r'(\d{2})(' + r'|'.join(MONTHS) + r')(\d{4})')


def cellify(values, output_worksheet, styles={}):
    for value in values:
        if type(value) in (tuple, list):
            value = ','.join(str(_value) for _value in value 
                if _value)
            if value == '':
                value = None
        
        cell = openpyxl.writer.dump_worksheet.WriteOnlyCell(
            output_worksheet, value)
        
        for attribute, style in styles.items():
            setattr(cell, attribute, style)
        else:
            yield cell


input_workbook = openpyxl.load_workbook('ISRID-2015-new.xlsx', read_only=True)
output_workbook = openpyxl.Workbook(write_only=True, optimized_write=True)
output_worksheet = output_workbook.create_sheet(title='survival-dataset')

base, step = 33, 14
base_style = {
    'font': openpyxl.styles.Font(name='Monospace', size=10)
}

with open('settings.yaml') as settings_file:
    settings = yaml.load(settings_file)
    weather.API_TOKENS = settings['tokens']
    weather._token_index = 0

output_worksheet.append(cellify([
    'Data Source', 
    'Key #', 
    'Incident Datetime', 
    'City', 
    'County', 
    'Subject Category', 
    'Age', 
    'Sex', 
    'Weight (kg)', 
    'Height (cm)', 
    'Subject Status', 
    'Total Incident Duration', 
    'IPP Coordinate (N/S degrees)', 
    'IPP Coordinate (E/W degrees)', 
    'High Temperature (degrees C)', 
    'Low Temperature (degrees C)', 
    'Wind Speed (km/h)', 
    'Snow (mm or boolean)', 
    'Rain (mm or boolean)'
], output_worksheet, base_style))

try:
    ignore = True
    for input_worksheet in input_workbook:
        util.log('Reading worksheet "{}" ... '.format(input_worksheet.title))
        for index, row in enumerate(input_worksheet.rows):
            if input_worksheet.title == 'NZ' and index == 83:
                print('IGNORE OFF')
                ignore = False
            
            if ignore:
                continue
            
            if index == 0:  # Heading with column names
                continue
            
            values = tuple(cell.value for cell in row)
            data_source, key = values[0], values[1]
            incident_datetime = values[4]
            location = values[8]
            city, county = values[9], values[10]
            category = values[20]
            age = tuple(values[base + i*step] for i in range(5))
            sex = tuple(values[base + i*step + 1] for i in range(5))
            weight = tuple(values[base + i*step + 3] for i in range(5))
            height = tuple(values[base + i*step + 4] for i in range(5))
            status = tuple(values[base + i*step + 13] for i in range(5))
            incident_duration = values[110]
            ipp = values[113]
            temp_max, temp_min, wind_speed = values[122], values[123], values[124]
            snow, rain = values[126], values[127]
            
            point = None
            
            if type(incident_datetime) is str:
                incident_datetime = incident_datetime.upper()
                result = DATETIME.search(incident_datetime)
                if result:
                    a, b, c = result.groups()
                    a, b, c = int(a), int(b), int(c)
                    if input_worksheet.title == 'US-VT':
                        year, month, day = a, b, c
                    elif input_worksheet.title == 'IE':
                        day, month, year = a, b, c + 2000
                    elif input_worksheet.title == 'PL':
                        day, month, year = a, b, c
                    elif input_worksheet.title in ('US NH', 'US GA', 'US WA', 
                            'US WY'):
                        year, month, day = c, a, b
                    elif input_worksheet.title == 'US UT':
                        month, day, year = a, b, c
                        if year < 100:
                            year += 2000
                    
                    try:
                        assert 1900 <= year <= 2015
                        incident_datetime = datetime.datetime(year, month, day)
                        util.log('Case {}-{} date: "{}" -> "{}"'.format(
                            input_worksheet.title, index, values[4], incident_datetime))
                    except (AssertionError, ValueError):
                        pass
                else:
                    result = DMY.search(incident_datetime)
                    if result:
                        day, month, year = result.groups()
                        day, month, year = int(day), MONTHS.index(month) + 1, \
                            int(year)
                        incident_datetime = datetime.datetime(year, month, day)
                        util.log('Case {}-{} date: "{}" -> "{}"'.format(
                            input_worksheet.title, index, values[4], incident_datetime))
                    else:
                        result = YMD.search(incident_datetime)
                        if result:
                            year, month, day = result.groups()
                            year, month, day = int(year), MONTHS.index(month) + 1, \
                                int(day)
                            incident_datetime = datetime.datetime(year, month, day)
                            util.log('Case {}-{} date: "{}" -> "{}"'.format(
                                input_worksheet.title, index, values[4], incident_datetime))
                        else:
                            pass  # Ignore
            
            if type(ipp) is str:
                try:
                    lat, lon = ipp.replace(' ', '').split(',')
                    lat, lon = float(lat), float(lon)
                    assert -90 <= lat <= 90 and -180 <= lon <= 180
                    point = geopy.Point(lat, lon)
                except ValueError:
                    pass
                except AssertionError:  # NZTM?
                    pass
            
            if not point and (bool(location) + bool(city) + bool(county) >= 2):
                query = '{} {} {}'.format(
                    location if location else '', 
                    city if city else '', 
                    county if county else ''
                ).strip()
                
                if input_worksheet.title != 'NZ':
                    point = geolocator.geocode(query, timeout=10)
                if point:
                    util.log('Case {}-{} IPP: "{}" -> "{}"'.format(
                        input_worksheet.title, index, query, point))
            
            if type(incident_datetime) is datetime.datetime and point:
                if not (temp_max and temp_min and wind_speed and rain and snow):
                    conditions = weather.get_conditions(incident_datetime, point)
                    print(conditions)
                    
                    if conditions['TMAX'] is not None and not temp_max:
                        temp_max = round(conditions['TMAX']/10, 3)
                        util.log('Case {}-{} Temp/H: "{}"'.format(
                            input_worksheet.title, index, temp_max))
                    
                    if conditions['TMIN'] is not None and not temp_min:
                        temp_min = round(conditions['TMIN']/10, 3)
                        util.log('Case {}-{} Temp/L: "{}"'.format(
                            input_worksheet.title, index, temp_min))
                    
                    if conditions['AWND'] is not None and not wind_speed:
                        wind_speed = round(3.6*conditions['AWND'], 3)
                        util.log('Case {}-{} Wind Speed: "{}"'.format(
                            input_worksheet.title, index, wind_speed))
                    
                    if type(snow) is str and 'NO' == snow.upper().strip():
                        snow = 0.0
                        util.log('Case {}-{} Snow: "{}"'.format(
                            input_worksheet.title, index, snow))
                    elif conditions['SNOW'] is not None and type(snow) not in (int, float):
                        snow = round(conditions['SNOW'], 3)
                        util.log('Case {}-{} Snow: "{}"'.format(
                            input_worksheet.title, index, snow))
                    
                    if type(rain) is str and 'NO' == rain.upper().strip():
                        rain = 0.0
                        util.log('Case {}-{} Rain: "{}"'.format(
                            input_worksheet.title, index, rain))
                    elif conditions['PRCP'] is not None and type(rain) not in (int, float):
                        rain = max(0.0, round(conditions['PRCP'] - (snow if snow else 0.0), 3))
                        util.log('Case {}-{} Rain: "{}"'.format(
                            input_worksheet.title, index, rain))
            
            output_worksheet.append(cellify([
                data_source, 
                key, 
                incident_datetime, 
                city, 
                county, 
                category, 
                age, 
                sex, 
                weight, 
                height, 
                status, 
                incident_duration, 
                point.latitude if point else None, 
                point.longitude if point else None, 
                temp_max, 
                temp_min, 
                wind_speed, 
                snow, 
                rain
            ], output_worksheet, base_style))
    
    output_workbook.save('ISRID-survival.xlsx')
    
    # input_workbook = openpyxl.load_workbook('ISRID-resolved.xlsx')
    # input_worksheet = input_workbook.active
except (KeyboardInterrupt, Exception) as e:
    print(e)
    output_workbook.save('ISRID-survival-dump.xlsx')
