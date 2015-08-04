#!/usr/bin/env python3
# SARbayes/ISRID/integrate3.py


import openpyxl


def cellify(values, output_worksheet, styles={}):
    for value in values:
        if type(value) in (tuple, list):
            value = ','.join(str(_value) for _value in value 
                if _value)
            if value == '':
                value = None
        
        cell = openpyxl.writer.dump_worksheet.WriteOnlyCell(
            output_worksheet, value)
        
        for attribute, style in styles.items():
            setattr(cell, attribute, style)
        else:
            yield cell

base_style = {
    'font': openpyxl.styles.Font(name='Monospace', size=10)
}


output_workbook = openpyxl.Workbook(optimized_write=True, write_only=True)
output_worksheet = output_workbook.create_sheet(title='survival-dataset')

input_workbook = openpyxl.load_workbook('ISRID-survival-2.xlsx', read_only=True)
input_worksheet = input_workbook.active

_input_workbook = openpyxl.load_workbook('ISRID-2015-new.xlsx', read_only=True)
_input_worksheet = None
for _worksheet in _input_workbook:
    if _worksheet.title == 'US OR':
        _input_worksheet = _worksheet
        break

assert _input_worksheet

__values = []
_count = 0
for _index, _row in enumerate(_input_worksheet.rows):
    if _index == 0:
        continue
    _values = list(cell.value for cell in _row)
    sex = [x for x in (_values[34], _values[48], _values[62], _values[90]) if x is not None]
    status = [x for x in (_values[46], _values[60], _values[74], _values[76], _values[102]) if x is not None]
    __values.append((sex, status))
    _count += 1

input_worksheet.max_row = 81707  # Important (defaults at 65536)

count = 0
for index, row in enumerate(input_worksheet.rows):
    values = list(cell.value for cell in row)
    
    if values[0] == 'US-OR' and values[1] is None:
        sex, status = __values[count]
        # values[6] = age
        values[7] = sex
        # values[8] = weight
        # values[9] = height
        values[10] = status
        count += 1
    
    output_worksheet.append(cellify(values, output_worksheet, base_style))

assert count == _count

with open('ISRID-NY.tab') as tab_file:
    text = tab_file.read().split('\n')
    for index, line in enumerate(text):
        if index >= 3 and line:
            key, status, age, sex, category, high_temp, low_temp, wind_speed, snow, rain = line.split('\t')
            if not key:
                key = None
            if not status:
                status = None
            if age:
                age = float(age)
            else:
                age = None
            if not sex:
                age = None
            if not category:
                category = None
            
            if high_temp:
                high_temp = float(high_temp)
            else:
                high_temp = None
            
            if low_temp:
                low_temp = float(low_temp)
            else:
                low_temp = None
            
            if wind_speed:
                wind_speed = float(wind_speed)
            else:
                wind_speed = None
            
            if snow:
                snow = float(snow)
            else:
                snow = None
            
            if rain:
                rain = float(rain)
            else:
                rain = None
            
            values = (
                'US-NY', 
                key, 
                None, 
                None, 
                None, 
                None, 
                category, 
                None, 
                age, 
                sex, 
                None, 
                None, 
                status, 
                None, 
                None, 
                None, 
                None, 
                high_temp, 
                low_temp, 
                wind_speed, 
                snow, 
                rain
            )
            
            output_worksheet.append(cellify(values, output_worksheet, base_style))

output_workbook.save('ISRID-survival.xlsx')
