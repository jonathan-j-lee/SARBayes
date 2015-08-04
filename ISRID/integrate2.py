#!/usr/bin/env python3
# SARbayes/ISRID/integrate2.py


import math
import openpyxl


def cellify(values, output_worksheet, styles={}):
    for value in values:
        if type(value) in (tuple, list):
            value = ','.join(str(_value) for _value in value 
                if _value)
            if value == '':
                value = None
        
        cell = openpyxl.writer.dump_worksheet.WriteOnlyCell(
            output_worksheet, value)
        
        for attribute, style in styles.items():
            setattr(cell, attribute, style)
        else:
            yield cell

base_style = {
    'font': openpyxl.styles.Font(name='Monospace', size=10)
}


output_workbook = openpyxl.Workbook(write_only=True, optimized_write=True)
output_worksheet = output_workbook.create_sheet(title='survival-dataset')
input_workbook = openpyxl.load_workbook('ISRID-resolved.xlsx', read_only=True)
input_worksheet = input_workbook.active

for index, row in enumerate(input_worksheet.rows):
    if index == 0:
        continue
    
    values = tuple(cell.value for cell in row)
    data_source = values[0]
    key = values[1]
    incident_datetime = values[3]
    city = values[4]
    county = values[5]
    category = values[11]
    age = values[15]
    sex = values[16]
    weight = values[19]
    height = values[20]
    incident_duration = values[31]
    lat = values[33]
    lon = values[34]
    temp_max = values[37]
    temp_min = values[38]
    wind_speed = values[39]
    snow = values[41]
    rain = values[42]
    status = values[46]
    
    output_worksheet.append(cellify([
        data_source, 
        key, 
        incident_datetime, 
        city, 
        county, 
        category, 
        age, 
        sex, 
        weight, 
        height, 
        status, 
        incident_duration, 
        lat, 
        lon, 
        temp_max, 
        temp_min, 
        wind_speed, 
        snow, 
        rain
    ], output_worksheet, base_style))

"""
import Orange
data = Orange.data.Table('ISRID-NY')
for instance in data:
    values = (
        'US-NY', 
        instance['Key #'].value if not math.isnan(instance['Key #']._value) else None, 
        None, 
        None, 
        None, 
        instance['Category'].value if not math.isnan(instance['Category']._value) else None, 
        instance['Age'].value if not math.isnan(instance['Age']._value) else None, 
        instance['Sex'].value if not math.isnan(instance['Sex']._value) else None, 
        None, 
        None, 
        instance['Status'].value if not math.isnan(instance['Status']._value) else None, 
        None, 
        None, 
        None, 
        instance['HighTemp'].value if not math.isnan(instance['HighTemp']._value) else None, 
        instance['LowTemp'].value if not math.isnan(instance['LowTemp']._value) else None, 
        instance['WindSpeed'].value if not math.isnan(instance['WindSpeed']._value) else None, 
        instance['Snow'].value if not math.isnan(instance['Snow']._value) else None, 
        instance['Rain'].value if not math.isnan(instance['Rain']._value) else None
    )
    output_worksheet.append(cellify(values, output_worksheet, base_style))
"""

input_workbook = openpyxl.load_workbook('ISRID-survival-dump-p1.xlsx', read_only=True)
input_worksheet = input_workbook.active

for index, row in enumerate(input_worksheet.rows):
    if index == 0:
        continue
    values = tuple(cell.value for cell in row)
    output_worksheet.append(cellify(values, output_worksheet, base_style))

input_workbook = openpyxl.load_workbook('ISRID-survival-dump-p2.xlsx', read_only=True)
input_worksheet = input_workbook.active

for index, row in enumerate(input_worksheet.rows):
    if index == 0:
        continue
    values = tuple(cell.value for cell in row)
    output_worksheet.append(cellify(values, output_worksheet, base_style))

input_workbook = openpyxl.load_workbook('ISRID-survival-dump-p3.xlsx', read_only=True)
input_worksheet = input_workbook.active

for index, row in enumerate(input_worksheet.rows):
    if index == 0:
        continue
    values = tuple(cell.value for cell in row)
    output_worksheet.append(cellify(values, output_worksheet, base_style))

input_workbook = openpyxl.load_workbook('ISRID-survival-dump-p4.xlsx', read_only=True)
input_worksheet = input_workbook.active

for index, row in enumerate(input_worksheet.rows):
    if index == 0:
        continue
    values = tuple(cell.value for cell in row)
    output_worksheet.append(cellify(values, output_worksheet, base_style))

input_workbook = openpyxl.load_workbook('ISRID-survival-dump-p5.xlsx')
input_worksheet = input_workbook.active

for index, row in enumerate(input_worksheet.rows):
    if index == 0:
        continue
    values = tuple(cell.value for cell in row)
    output_worksheet.append(cellify(values, output_worksheet, base_style))

output_workbook.save('ISRID-survival-2.xlsx')
