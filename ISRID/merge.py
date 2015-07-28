#!/usr/bin/env python3
# SARbayes/ISRID/merge.py


import formatting
import openpyxl
import util
import warnings
import weather
import yaml


def resolve_conflicts(row1, row2, output_worksheet, styles):
    for index, (cell1, cell2) in enumerate(zip(row1, row2)):
        value = None
        if cell1 != cell2 and cell1 is not None:
            if index in (37, 38, 39):
                value = float(cell1)
            elif index in (41, 42):
                if type(cell1) in (int, float):
                    value = float(cell1)
                elif 'NO' == cell1.strip().upper():
                    value = 0.0
                else:
                    try:
                        value = float(cell2)
                    except TypeError:
                        print(cell1)
                        value = cell1
            else:
                value = cell2
        else:
            value = cell2
        
        output_cell = openpyxl.writer.dump_worksheet.WriteOnlyCell(
            output_worksheet, value)
        
        for attribute, style in styles.items():
            setattr(output_cell, attribute, style)
        else:
            yield output_cell


def main():
    settings_filename = 'settings.yaml'
    input_filename1, input_filename2 = 'ISRIDclean.xlsx', 'ISRID-updated.xlsx'
    output_filename = 'ISRID-merged.xlsx'
    
    util.log('Reading settings file ... ')
    warnings.filterwarnings('ignore')
    
    with open(settings_filename) as settings_file:
        settings = yaml.load(settings_file)
        style_settings = settings.get('styles', dict())
        column_settings = settings.get('columns', dict())
        title, styles = settings.get('title', 'ISRID'), dict()
        weather.API_TOKENS = settings['tokens']
        
        util.log('Creating styles ... ')
        for attribute, arguments in style_settings.items():
            if attribute == 'font':
                style = openpyxl.styles.Font(**arguments)
            else:
                style = None
            styles[attribute] = style
    
    input_workbook1 = openpyxl.load_workbook(input_filename1, read_only=True)
    input_workbook2 = openpyxl.load_workbook(input_filename2, read_only=True)
    input_worksheet1, input_worksheet2 = (
        input_workbook1.active, input_workbook2.active)
    
    output_workbook = openpyxl.Workbook(write_only=True, optimized_write=True)
    output_worksheet = output_workbook.create_sheet(title=title)
    
    for row1, row2 in zip(input_worksheet1.rows, input_worksheet2.rows):
        row1 = list(cell.value for cell in row1)
        row2 = list(cell.value for cell in row2)
        output_worksheet.append(resolve_conflicts(row1, row2, 
            output_worksheet, styles))
    
    """
    input_workbook3 = openpyxl.load_workbook(input_filename3, read_only=True)
    for input_worksheet3 in input_workbook3:
        for index, row in enumerate(input_worksheet3.rows):
            if index == 0:
                continue
            row = list(cell.value for cell in row)
            if input_worksheet3.title == 'SDF':
                row = [
                    row[0],  # Data Source
                    row[1],  # Key#
                    row[6],  # Mission #
                    row[4],  # Incident Date
                    row[9],  # City
                    row[10],  # County
                    row[11],  # EcoRegion Domain
                    row[12],  # EcoRegion Division
                    row[13],  # Population Density
                    row[14],  # Terrain
                    row[3],  # Incident Type
                    row[20],  # Subject Category
                    row[21],  # Subject Sub-Category
                    row[23],  # Scenario
                    row[22],  # Subject Activity
                    row[33],  # Age
                    row[34],  # Sex
                    row[26],  # Number Lost
                ]
                print(row)
            else:
                pass
    """
    
    output_workbook.save(output_filename)
    util.log('Done.')


if __name__ == '__main__':
    main()
