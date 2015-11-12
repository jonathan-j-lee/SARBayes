#!/usr/bin/env python3

import datetime
import openpyxl
import weather
import yaml
import warnings
import geopy

warnings.filterwarnings('ignore')

with open('settings.yaml') as settings_file:
    settings = yaml.load(settings_file)
    weather.API_TOKENS = settings['tokens']

workbook_in = openpyxl.load_workbook('ISRID-2015-NY.xlsx', read_only=True)
worksheet_in = workbook_in.active

workbook_out = openpyxl.Workbook(write_only=True)
worksheet_out = workbook_out.create_sheet(title=worksheet_in.title)

for index, row in enumerate(worksheet_in.rows):
    values = [cell.value for cell in row]
    if index == 0:
        worksheet_out.append([openpyxl.writer.dump_worksheet.WriteOnlyCell(worksheet_out, value) for value in values])
        continue
    
    date, ipp = values[4], values[113]
    if type(date) is str:
        try:
            date = datetime.datetime.strptime(date, '%Y-%m-%d %H:%M:%S')
        except ValueError:
            date = datetime.datetime.strptime(date, '%Y-%m-%d')
        values[4] = date
    
    if ipp:
        lat, lon = ipp.split(', ')
        lat, lon = float(lat), float(lon)
        point = geopy.Point(lat, lon)
        conditions = weather.get_conditions(date, point)
        values[122] = round(conditions['TMAX']/10, 3) if conditions['TMAX'] is not None else None
        values[123] = round(conditions['TMIN']/10, 3) if conditions['TMIN'] is not None else None
        values[124] = round(conditions['AWND']*3.6, 3) if conditions['AWND'] is not None else None
        values[126] = round(conditions['SNOW'], 3) if conditions['SNOW'] is not None else None
        if conditions['PRCP'] is not None:
            if conditions['SNOW'] is not None:
                values[127] = round(max(conditions['PRCP'] - conditions['SNOW'], 0.0), 3)
            else:
                values[127] = round(conditions['PRCP'], 3)
        print(values[122:128], conditions, index)
    
    worksheet_out.append([openpyxl.writer.dump_worksheet.WriteOnlyCell(worksheet_out, value) for value in values])

workbook_out.save('ISRID-2015-NY-updated.xlsx')
